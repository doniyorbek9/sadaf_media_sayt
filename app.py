import os
import datetime
from flask import Flask, request, jsonify, send_file, render_template, abort
import requests
from openpyxl import Workbook, load_workbook

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

app = Flask(__name__, static_folder="static", template_folder="templates")

# ---------- Sozlamalar (Railway'da Environment Variables orqali beriladi) ----------
TELEGRAM_BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "")
ADMIN_CHAT_ID = os.environ.get("ADMIN_CHAT_ID", "")          # /README.md ga qarang - qanday olish
GROQ_API_KEY = os.environ.get("GROQ_API_KEY", "")
ADMIN_DOWNLOAD_KEY = os.environ.get("ADMIN_DOWNLOAD_KEY", "sadaf2026")  # /admin/orders?key=... uchun

ORDERS_FILE = os.path.join(os.path.dirname(__file__), "data", "orders.xlsx")
os.makedirs(os.path.dirname(ORDERS_FILE), exist_ok=True)

PACKAGES = {
    "standart": {
        "nomi": "Standart",
        "tavsif": "1 kunlik, 1 ta kamera",
        "narx": "700 000 so'm",
    },
    "premium": {
        "nomi": "Premium",
        "tavsif": "2 kunlik — 1-kun 1 ta kamera, 2-kun 2 ta kamera",
        "narx": "2 100 000 so'm",
    },
    "ultra": {
        "nomi": "Ultra",
        "tavsif": "2 kunlik — 1-kun 1 ta kamera, 2-kun 2 ta kamera + kran",
        "narx": "3 600 000 so'm",
    },
    "cinema_grand": {
        "nomi": "Cinema Grand",
        "tavsif": "3 kunlik — 1-kun 1 kamera, 2-kun 2 kamera + kran, 3-kun 1 kamera",
        "narx": "so'rov bo'yicha (hozircha narx belgilanmagan)",
    },
}

SERVICES = ["Nikoh to'yi", "Xatna (sunnat) to'yi", "Banket", "Tug'ilgan kun",
            "Xat-savod / Umra marosimlari", "Reklama roliklari", "Boshqa marosimlar"]

TEAM = [
    {"ism": "G'ofirjon", "rol": "Asoschi"},
    {"ism": "Diyorbek", "rol": "Videograf (jamoa a'zosi)"},
    {"ism": "Doniyorbek", "rol": "Videograf / Montajchi"},
    {"ism": "Shohruhbek", "rol": "Jamoa a'zosi"},
    {"ism": "Lochinbek", "rol": "Jamoa a'zosi"},
    {"ism": "Islombek", "rol": "Jamoa a'zosi"},
]

SYSTEM_PROMPT = f"""Sen "Sadaf Media" video studiyasining AI-yordamchisisan. Studiya to'y, xatna to'yi,
banket, tug'ilgan kun va boshqa marosimlar uchun professional videografiya xizmati ko'rsatadi.

Xizmatlar: {', '.join(SERVICES)}.

Paketlar:
- Standart: {PACKAGES['standart']['tavsif']}, narxi {PACKAGES['standart']['narx']}
- Premium: {PACKAGES['premium']['tavsif']}, narxi {PACKAGES['premium']['narx']}
- Ultra: {PACKAGES['ultra']['tavsif']}, narxi {PACKAGES['ultra']['narx']}
- Cinema Grand: {PACKAGES['cinema_grand']['tavsif']}, narxi {PACKAGES['cinema_grand']['narx']}

To'lov xizmat ko'rsatilgandan KEYIN qabul qilinadi (oldindan to'lov shart emas).

Vazifang: mijozga qisqa, samimiy va aniq javob berish, marosim turi va byudjetiga qarab mos
paketni tavsiya qilish, va agar mijoz band qilishga tayyor bo'lsa, saytdagi "Buyurtma" formasini
to'ldirishni taklif qilish. Javoblaring o'zbek tilida, qisqa (3-5 gap), do'stona va professional
bo'lsin. Narxlarni faqat yuqoridagi ro'yxatdan ol, o'zingdan narx to'qima."""


def ensure_orders_file():
    if not os.path.exists(ORDERS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Buyurtmalar"
        ws.append(["Sana/Vaqt", "Ism", "Telefon", "Marosim turi", "Marosim sanasi", "Paket", "Izoh"])
        wb.save(ORDERS_FILE)


def append_order(row):
    ensure_orders_file()
    wb = load_workbook(ORDERS_FILE)
    ws = wb["Buyurtmalar"]
    ws.append(row)
    wb.save(ORDERS_FILE)


TELEGRAM_API = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}"


def notify_telegram(text):
    if not TELEGRAM_BOT_TOKEN or not ADMIN_CHAT_ID:
        return False
    try:
        requests.post(f"{TELEGRAM_API}/sendMessage",
                      json={"chat_id": ADMIN_CHAT_ID, "text": text, "parse_mode": "HTML"}, timeout=10)
        return True
    except Exception as e:
        print("Telegram xato:", e)
        return False


def tg_send_message(chat_id, text, reply_markup=None):
    payload = {"chat_id": chat_id, "text": text, "parse_mode": "HTML"}
    if reply_markup:
        payload["reply_markup"] = reply_markup
    try:
        requests.post(f"{TELEGRAM_API}/sendMessage", json=payload, timeout=10)
    except Exception as e:
        print("Telegram xato:", e)


def tg_send_orders_file(chat_id):
    ensure_orders_file()
    try:
        with open(ORDERS_FILE, "rb") as f:
            requests.post(
                f"{TELEGRAM_API}/sendDocument",
                data={"chat_id": chat_id, "caption": "📊 Barcha buyurtmalar"},
                files={"document": ("sadaf-media-buyurtmalar.xlsx", f)},
                timeout=30,
            )
    except Exception as e:
        print("Telegram fayl yuborish xato:", e)


def tg_answer_callback(callback_query_id, text=None):
    try:
        payload = {"callback_query_id": callback_query_id}
        if text:
            payload["text"] = text
        requests.post(f"{TELEGRAM_API}/answerCallbackQuery", json=payload, timeout=10)
    except Exception as e:
        print("Telegram callback xato:", e)


ORDERS_BUTTON_MARKUP = {
    "inline_keyboard": [[{"text": "📊 Buyurtmalarni yuklab olish (Excel)", "callback_data": "get_orders"}]]
}


@app.route("/")
def index():
    return render_template("index.html", packages=PACKAGES, services=SERVICES, team=TEAM)


@app.route("/api/order", methods=["POST"])
def create_order():
    data = request.get_json(force=True, silent=True) or {}
    name = (data.get("name") or "").strip()
    phone = (data.get("phone") or "").strip()
    event_type = (data.get("event_type") or "").strip()
    event_date = (data.get("event_date") or "").strip()
    package = (data.get("package") or "").strip()
    note = (data.get("note") or "").strip()

    if not name or not phone:
        return jsonify({"ok": False, "error": "Ism va telefon raqami majburiy"}), 400

    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    append_order([now, name, phone, event_type, event_date, package, note])

    msg = (f"🎬 <b>Yangi buyurtma!</b>\n"
           f"Ism: {name}\nTelefon: {phone}\nMarosim: {event_type}\n"
           f"Sana: {event_date}\nPaket: {package}\nIzoh: {note or '-'}")
    notify_telegram(msg)

    return jsonify({"ok": True})


@app.route("/admin/orders")
def download_orders():
    key = request.args.get("key", "")
    if key != ADMIN_DOWNLOAD_KEY:
        abort(403)
    ensure_orders_file()
    return send_file(ORDERS_FILE, as_attachment=True, download_name="sadaf-media-buyurtmalar.xlsx")


@app.route("/telegram/webhook", methods=["POST"])
def telegram_webhook():
    update = request.get_json(force=True, silent=True) or {}

    # Oddiy matnli xabar (masalan /start)
    if "message" in update:
        msg = update["message"]
        chat_id = str(msg.get("chat", {}).get("id", ""))
        text = (msg.get("text") or "").strip()

        if chat_id != str(ADMIN_CHAT_ID):
            tg_send_message(chat_id, "Bu bot faqat Sadaf Media jamoasi uchun mo'ljallangan.")
            return jsonify({"ok": True})

        if text == "/start":
            tg_send_message(
                chat_id,
                "Assalomu alaykum! 🎬 Bu — Sadaf Media buyurtmalar boti.\n"
                "Saytdan tushgan har bir buyurtma shu yerga xabar bo'lib keladi.\n\n"
                "Pastdagi tugma orqali istalgan vaqt barcha buyurtmalarni Excel fayl "
                "sifatida yuklab olishingiz mumkin.",
                reply_markup=ORDERS_BUTTON_MARKUP,
            )
        else:
            tg_send_message(chat_id, "Buyurtmalarni yuklab olish uchun /start ni bosing.",
                             reply_markup=ORDERS_BUTTON_MARKUP)

    # Tugma bosilganda
    elif "callback_query" in update:
        cq = update["callback_query"]
        chat_id = str(cq.get("message", {}).get("chat", {}).get("id", ""))
        data = cq.get("data", "")

        if chat_id != str(ADMIN_CHAT_ID):
            tg_answer_callback(cq["id"], "Ruxsat yo'q")
            return jsonify({"ok": True})

        if data == "get_orders":
            tg_answer_callback(cq["id"], "Tayyorlanmoqda...")
            tg_send_orders_file(chat_id)

    return jsonify({"ok": True})


@app.route("/api/chat", methods=["POST"])
def chat():
    data = request.get_json(force=True, silent=True) or {}
    user_message = (data.get("message") or "").strip()
    history = data.get("history") or []  # [{role, content}, ...] oldingi xabarlar

    if not user_message:
        return jsonify({"ok": False, "error": "Xabar bo'sh"}), 400
    if not GROQ_API_KEY:
        return jsonify({"ok": False, "reply": "AI hozircha sozlanmagan. Iltimos, admin bilan bog'laning."}), 200

    messages = [{"role": "system", "content": SYSTEM_PROMPT}]
    for h in history[-10:]:
        if h.get("role") in ("user", "assistant") and h.get("content"):
            messages.append({"role": h["role"], "content": h["content"]})
    messages.append({"role": "user", "content": user_message})

    try:
        resp = requests.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={"Authorization": f"Bearer {GROQ_API_KEY}", "Content-Type": "application/json"},
            json={"model": "llama-3.3-70b-versatile", "messages": messages, "temperature": 0.5, "max_tokens": 400},
            timeout=20,
        )
        resp.raise_for_status()
        reply = resp.json()["choices"][0]["message"]["content"].strip()
        return jsonify({"ok": True, "reply": reply})
    except Exception as e:
        print("Groq xato:", e)
        return jsonify({"ok": False, "reply": "Kechirasiz, hozir javob bera olmadim. Birozdan so'ng qayta urinib ko'ring."}), 200


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
